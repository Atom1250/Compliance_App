"""Deterministic importer for curated regulatory source register files."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import inspect, select
from sqlalchemy.orm import Session

from apps.api.app.db.models import RegulatorySourceDocument

DEFAULT_XLSX_SHEETS = ("Master_Documents", "ESRS_Standards", "EU_Taxonomy_Acts")
CANONICAL_COLUMNS = (
    "record_id",
    "jurisdiction",
    "document_name",
    "document_type",
    "framework_level",
    "legal_reference",
    "issuing_body",
    "supervisory_authority",
    "official_source_url",
    "source_format",
    "language",
    "scope_applicability",
    "effective_date",
    "last_checked_date",
    "update_frequency",
    "version_identifier",
    "status",
    "keywords_tags",
    "notes_for_db_tagging",
    "source_sheets",
)
REQUIRED_COLUMNS = ("record_id", "jurisdiction", "document_name")


@dataclass(slots=True)
class ImportIssue:
    row_number: int
    sheet: str
    record_id: str
    field: str
    message: str


@dataclass(slots=True)
class ImportSummary:
    rows_seen: int = 0
    rows_deduped: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    invalid_rows: int = 0
    issues: list[ImportIssue] | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "rows_seen": self.rows_seen,
            "rows_deduped": self.rows_deduped,
            "inserted": self.inserted,
            "updated": self.updated,
            "skipped": self.skipped,
            "invalid_rows": self.invalid_rows,
        }


def _normalize_column_name(raw: str) -> str:
    value = raw.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def _normalize_tags(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    tokens = re.split(r"[|,;]+", text)
    deduped = sorted({re.sub(r"\s+", " ", token.strip()) for token in tokens if token.strip()})
    if not deduped:
        return None
    return "|".join(deduped)


def _parse_date(value: Any) -> date | None:
    text = _normalize_text(value)
    if text is None:
        return None
    if re.fullmatch(r"\d{4}", text):
        return date(int(text), 1, 1)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"invalid date: {text}")


def _coerce_optional_date(
    *,
    raw_value: Any,
    date_field: str,
    row_number: int,
    sheet: str,
    record_id: str,
    issues: list[ImportIssue],
) -> date | None:
    text = _normalize_text(raw_value)
    if text is None:
        return None
    try:
        return _parse_date(text)
    except ValueError:
        # Optional date fields should not fail ingestion; keep provenance in issues report.
        issues.append(
            ImportIssue(
                row_number=row_number,
                sheet=sheet,
                record_id=record_id,
                field=date_field,
                message=f"unparsed date retained as null: {text}",
            )
        )
        return None


def _normalize_url(value: Any) -> tuple[str | None, str | None]:
    text = _normalize_text(value)
    if text is None:
        return None, None
    lowered = text.lower()
    if not (lowered.startswith("http://") or lowered.startswith("https://")):
        return text, "URL must begin with http:// or https://"
    return text, None


def canonical_row_checksum(row: dict[str, Any]) -> str:
    """Return deterministic SHA256 checksum for one canonical source row."""
    canonical: dict[str, Any] = {}
    for key in sorted(CANONICAL_COLUMNS):
        value = row.get(key)
        if isinstance(value, date):
            canonical[key] = value.isoformat()
        elif isinstance(value, str):
            canonical[key] = value.strip()
        else:
            canonical[key] = value
    payload = json.dumps(canonical, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _normalize_row(
    *,
    row: dict[str, Any],
    row_number: int,
    sheet: str,
    issues: list[ImportIssue],
) -> dict[str, Any] | None:
    normalized = {column: None for column in CANONICAL_COLUMNS}
    for key, value in row.items():
        if key not in normalized:
            continue
        normalized[key] = _normalize_text(value)

    normalized["keywords_tags"] = _normalize_tags(normalized.get("keywords_tags"))

    missing_required = [field for field in REQUIRED_COLUMNS if not normalized.get(field)]
    if missing_required:
        issues.append(
            ImportIssue(
                row_number=row_number,
                sheet=sheet,
                record_id=normalized.get("record_id") or "",
                field="|".join(missing_required),
                message="required field(s) missing",
            )
        )
        return None

    record_id = normalized["record_id"]
    for date_field in ("effective_date", "last_checked_date"):
        normalized[date_field] = _coerce_optional_date(
            raw_value=normalized.get(date_field),
            date_field=date_field,
            row_number=row_number,
            sheet=sheet,
            record_id=record_id,
            issues=issues,
        )

    normalized["official_source_url"], url_issue = _normalize_url(
        normalized.get("official_source_url")
    )
    if url_issue is not None:
        issues.append(
            ImportIssue(
                row_number=row_number,
                sheet=sheet,
                record_id=record_id,
                field="official_source_url",
                message=url_issue,
            )
        )
    return normalized


def _merge_rows(base: dict[str, Any], incoming: dict[str, Any], *, sheet: str) -> dict[str, Any]:
    merged = dict(base)
    for column in CANONICAL_COLUMNS:
        if column == "source_sheets":
            continue
        current = merged.get(column)
        candidate = incoming.get(column)
        if current in (None, "") and candidate not in (None, ""):
            merged[column] = candidate
    prior_sheets = (
        set((base.get("source_sheets") or "").split("|")) if base.get("source_sheets") else set()
    )
    incoming_sheets = (
        set((incoming.get("source_sheets") or "").split("|"))
        if incoming.get("source_sheets")
        else set()
    )
    combined = sorted(
        sheet_name for sheet_name in (prior_sheets | incoming_sheets | {sheet}) if sheet_name
    )
    merged["source_sheets"] = "|".join(combined) if combined else None
    return merged


def _normalized_csv_rows(file_path: Path) -> tuple[list[dict[str, Any]], set[str]]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV has no header row: {file_path}")
        columns = {_normalize_column_name(field): field for field in reader.fieldnames}
        missing = [field for field in REQUIRED_COLUMNS if field not in columns]
        if missing:
            raise ValueError(f"CSV missing required columns: {', '.join(missing)}")
        rows = []
        for source_row in reader:
            normalized_row = {
                _normalize_column_name(column): value
                for column, value in source_row.items()
                if column
            }
            normalized_row["__sheet"] = "csv"
            rows.append(normalized_row)
        return rows, {"csv"}


def _normalized_xlsx_rows(
    file_path: Path, sheets: tuple[str, ...] | None
) -> tuple[list[dict[str, Any]], set[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValueError(
            "XLSX import requires openpyxl. Install it in the runtime environment."
        ) from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    available = set(workbook.sheetnames)
    selected = (
        list(sheets) if sheets else [name for name in DEFAULT_XLSX_SHEETS if name in available]
    )
    if not selected:
        selected = sorted(available)

    rows: list[dict[str, Any]] = []
    used_sheets: set[str] = set()
    for sheet_name in selected:
        if sheet_name not in available:
            continue
        used_sheets.add(sheet_name)
        worksheet = workbook[sheet_name]
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iter)
        except StopIteration:
            continue
        headers = [_normalize_column_name(str(value or "")) for value in raw_headers]
        missing = [field for field in REQUIRED_COLUMNS if field not in headers]
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' missing required columns: {', '.join(missing)}")
        for values in row_iter:
            if values is None:
                continue
            row_map = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
            row_map["__sheet"] = sheet_name
            rows.append(row_map)
    return rows, used_sheets


def _load_source_rows(
    file_path: Path, sheets: tuple[str, ...] | None
) -> tuple[list[dict[str, Any]], set[str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _normalized_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _normalized_xlsx_rows(file_path, sheets)
    raise ValueError(f"Unsupported file type: {file_path.suffix}")


def write_issues_report(path: Path, issues: list[ImportIssue]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["row_number", "sheet", "record_id", "field", "message"])
        for issue in issues:
            writer.writerow(
                [issue.row_number, issue.sheet, issue.record_id, issue.field, issue.message]
            )


def import_regulatory_sources(
    db: Session,
    *,
    file_path: Path,
    sheets: tuple[str, ...] | None = None,
    jurisdiction: str | None = None,
    dry_run: bool = False,
    issues_out: Path | None = None,
) -> ImportSummary:
    """Import one curated source register file into regulatory_source_document."""
    rows, _ = _load_source_rows(file_path, sheets=sheets)
    issues: list[ImportIssue] = []
    summary = ImportSummary(issues=issues)

    deduped: dict[str, dict[str, Any]] = {}

    for row_number, row in enumerate(rows, start=2):
        summary.rows_seen += 1
        row_sheet = str(row.pop("__sheet", "csv") or "csv")
        normalized = _normalize_row(row=row, row_number=row_number, sheet=row_sheet, issues=issues)
        if normalized is None:
            summary.invalid_rows += 1
            continue
        if jurisdiction and normalized["jurisdiction"] != jurisdiction:
            continue
        normalized["source_sheets"] = normalized.get("source_sheets") or row_sheet
        existing = deduped.get(normalized["record_id"])
        if existing is None:
            deduped[normalized["record_id"]] = normalized
            continue
        deduped[normalized["record_id"]] = _merge_rows(existing, normalized, sheet=row_sheet)

    summary.rows_deduped = len(deduped)

    for record_id in sorted(deduped):
        row = deduped[record_id]
        row["row_checksum"] = canonical_row_checksum(row)
        row["raw_row_json"] = {
            key: (value.isoformat() if isinstance(value, date) else value)
            for key, value in row.items()
            if key in CANONICAL_COLUMNS
        }

    if dry_run:
        if issues_out is not None:
            write_issues_report(issues_out, issues)
        return summary

    if not inspect(db.get_bind()).has_table("regulatory_source_document"):
        raise ValueError(
            "Table regulatory_source_document does not exist. "
            "Apply migrations (alembic upgrade head) then retry."
        )

    existing_records = {
        record.record_id: record
        for record in db.scalars(
            select(RegulatorySourceDocument).where(
                RegulatorySourceDocument.record_id.in_(sorted(deduped))
            )
        ).all()
    }

    now_utc = datetime.now(UTC)
    updatable_fields = [
        "jurisdiction",
        "document_name",
        "document_type",
        "framework_level",
        "legal_reference",
        "issuing_body",
        "supervisory_authority",
        "official_source_url",
        "source_format",
        "language",
        "scope_applicability",
        "effective_date",
        "last_checked_date",
        "update_frequency",
        "version_identifier",
        "status",
        "keywords_tags",
        "notes_for_db_tagging",
        "source_sheets",
    ]

    for record_id in sorted(deduped):
        row = deduped[record_id]
        checksum = row["row_checksum"]
        existing = existing_records.get(record_id)
        if existing is None:
            db.add(
                RegulatorySourceDocument(
                    record_id=record_id,
                    jurisdiction=row["jurisdiction"],
                    document_name=row["document_name"],
                    document_type=row["document_type"],
                    framework_level=row["framework_level"],
                    legal_reference=row["legal_reference"],
                    issuing_body=row["issuing_body"],
                    supervisory_authority=row["supervisory_authority"],
                    official_source_url=row["official_source_url"],
                    source_format=row["source_format"],
                    language=row["language"],
                    scope_applicability=row["scope_applicability"],
                    effective_date=row["effective_date"],
                    last_checked_date=row["last_checked_date"],
                    update_frequency=row["update_frequency"],
                    version_identifier=row["version_identifier"],
                    status=row["status"],
                    keywords_tags=row["keywords_tags"],
                    notes_for_db_tagging=row["notes_for_db_tagging"],
                    source_sheets=row["source_sheets"],
                    row_checksum=checksum,
                    raw_row_json=row["raw_row_json"],
                    created_at=now_utc,
                    updated_at=now_utc,
                )
            )
            summary.inserted += 1
            continue

        if existing.row_checksum == checksum:
            summary.skipped += 1
            continue

        changed = False
        for field in updatable_fields:
            incoming_value = row[field]
            if incoming_value in (None, ""):
                continue
            if getattr(existing, field) != incoming_value:
                setattr(existing, field, incoming_value)
                changed = True
        if existing.row_checksum != checksum:
            existing.row_checksum = checksum
            changed = True
        if existing.raw_row_json != row["raw_row_json"]:
            existing.raw_row_json = row["raw_row_json"]
            changed = True
        if changed:
            existing.updated_at = now_utc
            summary.updated += 1
        else:
            summary.skipped += 1

    db.commit()
    if issues_out is not None:
        write_issues_report(issues_out, issues)
    return summary
